# src/analytics/peer.py

import os
import sqlite3
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from ..config import DB_PATH, OUTPUT_DIR

# Output paths
PEER_OUT_PATH = os.path.join(OUTPUT_DIR, "peer_comparison.xlsx")
RADAR_DIR = os.path.join(OUTPUT_DIR, "../reports/radar_charts")

def setup_peer_tables(conn):
    """Initializes the output table in SQLite."""
    cursor = conn.cursor()
    cursor.execute("DROP TABLE IF EXISTS peer_percentiles;")
    cursor.execute("""
    CREATE TABLE peer_percentiles (
        company_id TEXT,
        peer_group_name TEXT,
        metric TEXT,
        value REAL,
        percentile_rank REAL,
        year TEXT,
        PRIMARY KEY (company_id, metric, year)
    );
    """)
    conn.commit()

def generate_radar_chart(comp_id, peer_name, company_metrics, peer_averages):
    """Plots a filled polar polygon with peer group benchmarks."""
    categories = list(company_metrics.keys())
    N = len(categories)

    # Calculate angles for each axis
    angles = [n / float(N) * 2 * np.pi for n in range(N)]
    angles += angles[:1]  # complete the loop

    values = list(company_metrics.values())
    values += values[:1]

    avg_values = list(peer_averages.values())
    avg_values += avg_values[:1]

    fig, ax = plt.subplots(figsize=(6, 6), subplot_kw=dict(polar=True))
    
    # Draw one axis per variable and add labels
    plt.xticks(angles[:-1], categories, color='grey', size=8)

    # Plot company statistics
    ax.plot(angles, values, linewidth=2, linestyle='solid', label=comp_id, color="#007BFF")
    ax.fill(angles, values, 'b', alpha=0.1)

    # Overlay peer averages
    ax.plot(angles, avg_values, linewidth=1.5, linestyle='dashed', label=f"{peer_name} Avg", color="#DC3545")

    plt.title(f"Peer KPI Comparison - {comp_id}", size=11, color='#333333', y=1.1)
    plt.legend(loc='upper right', bbox_to_anchor=(0.1, 0.1))
    
    os.makedirs(RADAR_DIR, exist_ok=True)
    chart_path = os.path.join(RADAR_DIR, f"{comp_id}_radar.png")
    plt.savefig(chart_path, bbox_inches='tight', dpi=150)
    plt.close()

def compute_peer_percentiles():
    conn = sqlite3.connect(DB_PATH)
    setup_peer_tables(conn)

    # Load financial ratio data
    ratios = pd.read_sql_query("SELECT * FROM financial_ratios", conn)
    
    # Dynamically resolve available columns in sectors table
    cursor = conn.cursor()
    cursor.execute("PRAGMA table_info(sectors);")
    columns = [col[1] for col in cursor.fetchall()]
    print(f"🔍 Found columns in 'sectors' table: {columns}")

    # Determine best column to use for peer grouping
    if "peer_group" in columns:
        peer_col = "peer_group"
    elif "sub_sector" in columns:
        peer_col = "sub_sector"
    elif "broad_sector" in columns:
        peer_col = "broad_sector"
    else:
        peer_col = columns[1] if len(columns) > 1 else "company_id"

    print(f"🎯 Using column '{peer_col}' as 'peer_group'")
    peer_mapping = pd.read_sql_query(f"SELECT company_id, {peer_col} AS peer_group FROM sectors", conn)

    # Safely merge mapping to establish the complete base DataFrame 'm_df'
    m_df = pd.merge(ratios, peer_mapping, on="company_id", how="left")
    m_df['peer_group'] = m_df['peer_group'].fillna("No peer group assigned")

    metrics_to_rank = [
        "return_on_equity_pct", "net_profit_margin_pct", "debt_to_equity",
        "interest_coverage", "asset_turnover", "free_cash_flow_cr"
    ]

    records_to_insert = []
    wb = Workbook()
    wb.remove(wb.active)  # Clear default sheet

    # Styling definitions
    green_fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
    red_fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
    header_fill = PatternFill(start_color="17A2B8", end_color="17A2B8", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    # Group by the dynamically evaluated peer group column
    for group_name, group_df in m_df.groupby("peer_group"):
        if group_name == "No peer group assigned":
            continue

        # Setup sheet for peer group comparison
        ws = wb.create_sheet(title=str(group_name)[:30])
        headers = ["company_id", "year"] + metrics_to_rank + [f"{m}_pct_rank" for m in metrics_to_rank]
        ws.append(headers)
        
        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill

        # Compute averages for radar charts
        peer_averages = {m: group_df[m].mean() if not pd.isna(group_df[m].mean()) else 0.0 for m in metrics_to_rank}

        # Calculate Percentile Rankings
        for _, row in group_df.iterrows():
            row_dict = {"company_id": row["company_id"], "year": row["year"]}
            company_metrics = {}

            for metric in metrics_to_rank:
                val = row[metric]
                
                # Compute ranking within the peer group for the current year
                values_list = group_df[group_df['year'] == row['year']][metric].dropna().tolist()
                
                if len(values_list) > 1 and not pd.isna(val):
                    pct_rank = sum(1 for v in values_list if v <= val) / len(values_list)
                    
                    # Invert debt rank (lower leverage = better rank)
                    if metric == "debt_to_equity":
                        pct_rank = 1.0 - pct_rank
                else:
                    pct_rank = 0.5  # Default middle rank

                row_dict[metric] = val
                row_dict[f"{metric}_pct_rank"] = round(pct_rank * 100.0, 1)
                
                # Collect coordinates for Radar Chart
                company_metrics[metric] = pct_rank * 100.0

                records_to_insert.append({
                    "company_id": row["company_id"],
                    "peer_group_name": group_name,
                    "metric": metric,
                    "value": val,
                    "percentile_rank": round(pct_rank * 100.0, 2),
                    "year": row["year"]
                })

            # Append structured row data to comparison sheet
            row_vals = [row_dict[h] for h in headers]
            ws.append(row_vals)
            curr_row = ws.max_row

            # Apply cell colors to percentile ranks
            for col_idx, col_name in enumerate(headers, 1):
                if "_pct_rank" in col_name:
                    cell = ws.cell(row=curr_row, column=col_idx)
                    score = cell.value or 0
                    
                    if score >= 75.0:
                        cell.fill = green_fill
                    elif score >= 25.0:
                        cell.fill = yellow_fill
                    else:
                        cell.fill = red_fill

            # Generate radar charts
            generate_radar_chart(row["company_id"], group_name, company_metrics, peer_averages)

    # Save SQL data
    df_db = pd.DataFrame(records_to_insert)
    if not df_db.empty:
        df_db.to_sql("peer_percentiles", conn, if_exists="append", index=False)
        conn.commit()

    wb.save(PEER_OUT_PATH)
    conn.close()
    print(f"📊 Peer Comparison metrics saved! Excel report exported to '{PEER_OUT_PATH}'")

if __name__ == "__main__":
    compute_peer_percentiles()