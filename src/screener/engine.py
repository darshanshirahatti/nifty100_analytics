# src/screener/engine.py

import os
import yaml
import sqlite3
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment

# Robust directory paths
BASE_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", ".."))
DB_PATH = os.path.join(BASE_DIR, "data", "nifty100.db")
CONFIG_PATH = os.path.join(BASE_DIR, "config", "screener_config.yaml")
OUTPUT_PATH = os.path.join(BASE_DIR, "output", "screener_output.xlsx")

def winsorise_and_scale(series, p_low=10, p_high=90):
    """Caps extreme values at specified percentiles and scales them linearly between 0 and 100."""
    series = pd.to_numeric(series, errors='coerce').fillna(0.0)
    low_val = np.percentile(series, p_low)
    high_val = np.percentile(series, p_high)
    
    # Winsorise/Cap values
    capped = np.clip(series, low_val, high_val)
    
    # Linear scaling to 0-100
    denom = (high_val - low_val)
    if denom == 0:
        return capped * 0.0
    return ((capped - low_val) / denom) * 100.0

def compute_composite_scores(df, financials_list):
    """
    Computes winsorised Quality Scores:
    35% Profitability + 30% Cash Quality + 20% Growth + 15% Leverage
    """
    # 1. Profitability (ROE + ROCE + NPM)
    roe_s = winsorise_and_scale(df['return_on_equity_pct'])
    npm_s = winsorise_and_scale(df['net_profit_margin_pct'])
    prof_score = roe_s * 0.5 + npm_s * 0.5

    # 2. Cash Quality (CFO/PAT proxy, positive check)
    fcf_pos = df['free_cash_flow_cr'].apply(lambda x: 100.0 if (x and x > 0) else 0.0)
    cash_score = fcf_pos

    # 3. Growth (Rev CAGR)
    rev_g = winsorise_and_scale(df['revenue_cagr_5yr'])
    growth_score = rev_g

    # 4. Leverage
    de_s = df['debt_to_equity'].fillna(0.0)
    # Inverse: Lower is better. Scale and subtract from 100.
    de_scaled = 100.0 - winsorise_and_scale(de_s)
    
    # Set maximum score for Financials
    de_scaled = df.apply(lambda row: 100.0 if row['company_id'] in financials_list else de_scaled[row.name], axis=1)

    # Sum everything according to weighted index guidelines
    df['composite_quality_score'] = (prof_score * 0.35) + (cash_score * 0.30) + (growth_score * 0.20) + (de_scaled * 0.15)
    df['composite_quality_score'] = df['composite_quality_score'].round(2)
    return df

def run_screener():
    # Load settings
    with open(CONFIG_PATH, "r") as f:
        config = yaml.safe_load(f)

    conn = sqlite3.connect(DB_PATH)
    
    # Fetch base details
    df = pd.read_sql_query("SELECT * FROM financial_ratios", conn)
    
    try:
        sectors = pd.read_sql_query("SELECT company_id, broad_sector FROM sectors", conn)
        financial_sector_companies = set(sectors[sectors['broad_sector'].str.upper() == 'FINANCIALS']['company_id'])
    except Exception:
        financial_sector_companies = set()

    conn.close()

    # Apply Quality Index calculations
    df = compute_composite_scores(df, financial_sector_companies)

    # Initialize Excel writer
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet

    # Stylers
    green_fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
    red_fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="343A40", end_color="343A40", fill_type="solid")

    for preset_name, spec in config['presets'].items():
        metrics = spec['metrics']
        filtered_df = df.copy()

        # Apply each condition
        for col, rules in metrics.items():
            if col not in filtered_df.columns:
                continue

            # Convert column to numeric safely
            filtered_df[col] = pd.to_numeric(filtered_df[col], errors='coerce')

            # Minimum checks
            if 'min' in rules:
                val = rules['min']
                # Treat infinity and edge-cases carefully (like ICR Label conversions)
                filtered_df = filtered_df[filtered_df[col].fillna(float('inf')) >= val]
                
            # Maximum checks
            if 'max' in rules:
                val = rules['max']
                if col == 'debt_to_equity':
                    # Skip companies in Financials sector from strict D/E limits
                    filtered_df = filtered_df[
                        (filtered_df[col] <= val) | 
                        (filtered_df['company_id'].isin(financial_sector_companies))
                    ]
                else:
                    filtered_df = filtered_df[filtered_df[col].fillna(0.0) <= val]

        # Order by computed score
        filtered_df = filtered_df.sort_values(by="composite_quality_score", ascending=False)

        # Create sheet
        ws = wb.create_sheet(title=preset_name.replace("_", " ").title()[:30])
        
        # Write Headers
        headers = list(filtered_df.columns)
        ws.append(headers)
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        # Write Data & color-code cells based on preset requirements
        for _, row in filtered_df.iterrows():
            row_data = [row[h] for h in headers]
            ws.append(row_data)
            curr_row = ws.max_row

            # Apply cell-by-cell conditional formatting
            for col_idx, col_name in enumerate(headers, 1):
                cell = ws.cell(row=curr_row, column=col_idx)
                
                if col_name in metrics:
                    rules = metrics[col_name]
                    cell_val = row[col_name]
                    
                    if pd.isna(cell_val):
                        continue
                        
                    is_passing = True
                    if 'min' in rules and cell_val < rules['min']:
                        is_passing = False
                    if 'max' in rules and cell_val > rules['max']:
                        if not (col_name == 'debt_to_equity' and row['company_id'] in financial_sector_companies):
                            is_passing = False
                            
                    cell.fill = green_fill if is_passing else red_fill

    wb.save(OUTPUT_PATH)
    print(f"🎯 Screener calculations completed! Results saved to '{OUTPUT_PATH}'")

if __name__ == "__main__":
    run_screener()